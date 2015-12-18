#!/usr/bin/env python3
# encoding: utf-8

import os
from base import *
from models import *
from openpyxl import Workbook
from openpyxl import load_workbook
import uuid
import random
import json
import tempfile

def uuid64bit():
    # return uuid.uuid4().int >> 64
    return random.getrandbits(32)

class account_meta_handler(base_handler):
    def get(self, account_id):
        meta = BillMeta.objects(id = account_id).first()
        res = {}
        res['id'] = account_id
        res['pages'] =  acount_pages
        res['col_names'] = []
        for col in  meta.col_names:
            res['col_names'].append(col)
        print(res)
        self.write(json.dumps(res))

    def post(self, account_id):
        bill_meta = BillMeta.objects(id = account_id).first()
        bill_meta['col_names']

class account_handler(base_handler):
    def get(self, accout_id):
        my_account = Bill.objects(id = int(account_id)).first()
        xls = my_account.xls.read()
        with tempfile.NamedTemporaryFile(delete=False) as fp:
            name = fp.name
            fp.write(xls)
            fp.close()
            os.rename(name, name + '.xlsx')
            excel = load_workbook(name + '.xlsx')
            worksheet = excel['test']
            to_sheet = []
            for row in worksheet.rows:
                to_row = []
                for cell in row:
                    to_cell = {
                            'coordinate' : cell.coordinate,
                            'value' : cell.value and cell.value or 0
                    }
                    to_row.append(to_cell)
                to_sheet.append(to_row)
            self.write(json.dumps(to_sheet))

    def post(self, account_id):
        coordinate = self.get_json_argument('coordinate')
        value = self.get_json_argument('value')
        bill = Bill.objects(id = int(account_id)).first()
        xls = bill.xls.read()
        with tempfile.NamedTemporaryFile(delete=False) as fp:
            name = fp.name
            fp.write(xls)
            fp.close()
            os.rename(name, name + '.xlsx')
            excel = load_workbook(name + '.xlsx')
            worksheet = excel['test']
            worksheet[coordinate] = value
            excel.save(name + '.xlsx')
            new_excel = open(name + '.xlsx', 'rb')
            bill.xls.replace(new_excel)
            bill.save()
            self.write('')

class add_column_handler(base_handler):
    def post(self, account_id):
        name  = self.get_json_argument('name')
        defaultVal = self.get_json_argument('defaultVal')
        bill = Bill.objects(id = int(account_id)).first()
        xls = bill.xls.read()
        with tempfile.NamedTemporaryFile(delete=False) as fp:
            name = fp.name
            fp.write(xls)
            fp.close()
            os.rename(name, name + '.xlsx')
            excel = load_workbook(name + '.xlsx')
            worksheet = excel['test']
            cols = len(worksheet.columns) + 1
            rows = len(worksheet.rows)
            for i in range(1, rows + 1):
                worksheet.cell(row = i, column = cols).value = defaultVal
            excel.save(name + '.xlsx')
            new_excel = open(name + '.xlsx', 'rb')
            bill.xls.replace(new_excel)
            bill.save()
            self.write('')
        pass

class account_test_handler(base_handler):
    def get(self):
        wb = Workbook()
        ws = wb.create_sheet()
        ws.title = 'test'
        ws['A4'] = 4
        bill = Bill(id = 1, name = 'test', obj = 'test')
        with tempfile.NamedTemporaryFile(delete = False) as fp:
            name = fp.name
            fp.close()
            wb.save(name)
            data = open(name, 'rb')
            print(data)
            bill.xls.put(data)
            bill.save()
            os.unlink(name)

class account_tree_handler(base_handler):
    def get(self):
        res = []
        for cat in Cat.objects:
            item = {
                    'id' : cat.id,
                    'name' : cat.name,
            }
            item['children'] = []
            for aset in cat.cats:
                subitem = {
                        'id' : aset.id,
                        'name' : aset.name,
                    }
                subitem['children'] = []
                for bill in aset.sets:
                    abill = {
                            'id' : bill.id,
                            'name' : bill.name
                        }
                    subitem['children'].append(abill)
                item['children'].append(subitem)
            res.append(item)
        print(res)
        self.write(json.dumps(res))

class account_add_handler(base_handler):
    def post(self):
        account_name = self.get_json_argument('name')
        account_type = self.get_json_argument('type')
        account_year = self.get_json_argument('year')
        bill = Bill(id = uuid64bit(), name = account_name, obj = '')
        wb = Workbook()
        ws = wb.create_sheet()
        ws.title = 'test'
        with tempfile.NamedTemporaryFile(delete = False) as fp:
            name = fp.name
            fp.close()
            wb.save(name)
            data = open(name, 'rb')
            bill.xls.put(data)
            os.unlink(name)
        bill.save()
        if Cat.objects(name = account_year).first():
            cat =  Cat.objects(name = account_year).first()
            isFound = False
            for aset in cat.cats:
                if aset.name == account_type:
                    aset.sets.append(bill)
                    aset.save()
                    isFound = True
                    break
            if not isFound:
                myset = Set(id = uuid64bit(), name = account_type, sets = [bill])
                myset.save()
                cat.cats.append(myset)
        else:
            myset = Set(id = uuid64bit(), name = account_type, sets = [bill])
            myset.save()
            cat = Cat(id = uuid64bit(), name = account_year, cats = [myset])
        cat.save()
        self.write('')

class test_tree_handler(base_handler):
    def get(self):
        bill = Bill(id = uuid64bit(), name = 'test', obj = '123')
        bill.save()
        aset = Set(id = uuid64bit(), name = 'test', sets = [bill])
        aset.save()
        cat = Cat(id = uuid64bit(), name = 'test', cats = [aset])
        cat.save()
